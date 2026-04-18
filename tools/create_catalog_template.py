from openpyxl import Workbook


def create_catalog(filepath="data/equipment_catalog.xlsx"):
    wb = Workbook()

    # Лист 1: типы оборудования
    ws1 = wb.active
    ws1.title = "equipment_types"
    ws1.append([
        "type_id",
        "type_name",
        "energy_source",
        "default_qy_kw",
        "ka_w_per_kw",
        "notes"
    ])
    ws1.append(["josper_grill", "Гриль-печь JOSPER", "electric", 6.0, 720, "Пример"])
    ws1.append(["combi_oven", "Пароконвектомат", "electric", 11.7, 120, "Пример"])
    ws1.append(["induction_hob", "Плита индукционная", "electric", 21.0, 200, "Пример"])
    ws1.append(["deep_fryer", "Фритюрница", "electric", 15.0, 90, "Пример"])

    # Лист 2: категории помещений и Kз
    ws2 = wb.create_sheet("room_categories")
    ws2.append([
        "room_category",
        "kz_default",
        "notes"
    ])
    ws2.append(["Горячий цех", 0.7, "Типовое значение"])
    ws2.append(["Кондитерский цех", 0.6, "Пример"])
    ws2.append(["Мясной цех", 0.5, "Пример"])

    # Лист 3: положение и коэффициент r
    ws3 = wb.create_sheet("position_coefficients")
    ws3.append([
        "position_name",
        "r_value",
        "notes"
    ])
    ws3.append(["Свободно стоящее", 1.0, "Таблица положения"])
    ws3.append(["У стены", 0.63, "Минимум из таблицы"])
    ws3.append(["В углу", 0.4, "Таблица положения"])

    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(filepath)
    print(f"Каталог создан: {filepath}")


if __name__ == "__main__":
    create_catalog()