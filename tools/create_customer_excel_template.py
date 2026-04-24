from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment


def create_template(filepath="data/customer_input_template.xlsx"):
    wb = Workbook()

    ws = wb.active
    ws.title = "Исходные данные"

    headers = [
        "Наименование помещения",
        "Оборудование",
        "Ширина",
        "Глубина",
    ]

    ws.append(headers)

    examples = [
        ["Кухня 1", "Плита", 700, 400],
        ["Кухня 1", "Пароконвектомат", 750, 783],
        ["Кухня 1", "Фритюрница", 400, 700],
        ["Кухня 2", "Печь для пиццы", 1000, 900],
    ]

    for row in examples:
        ws.append(row)

    # Ширины колонок
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    # Стили
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Таблица Excel
    table = Table(displayName="tbl_customer_input", ref="A1:D5")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Комментарии к заголовкам
    ws["A1"].comment = Comment("Например: Кухня 1, Кухня 2, Пицца", "ChatGPT")
    ws["B1"].comment = Comment("Пишите как на плане или в спецификации. Тип оборудования выберет инженер.", "ChatGPT")
    ws["C1"].comment = Comment("Ширина оборудования в мм. Только число.", "ChatGPT")
    ws["D1"].comment = Comment("Глубина оборудования в мм. Только число.", "ChatGPT")

    # Инструкция
    ws2 = wb.create_sheet("Инструкция")
    ws2["A1"] = "Инструкция по заполнению"
    ws2["A1"].font = Font(bold=True, size=14)

    instruction = [
        "1. Заполняйте только лист 'Исходные данные'.",
        "2. Каждая строка = одна единица оборудования.",
        "3. Не объединяйте ячейки.",
        "4. Не удаляйте заголовки колонок.",
        "5. Наименование помещения пишите одинаково для оборудования из одного помещения.",
        "6. Ширина и глубина указываются в миллиметрах.",
        "7. Если размер неизвестен — поставьте 0.",
        "8. Тип оборудования, категорию помещения и расчётные коэффициенты выберет инженер.",
    ]

    for i, text in enumerate(instruction, start=3):
        ws2[f"A{i}"] = text

    ws2.column_dimensions["A"].width = 100

    wb.save(filepath)
    print(f"Шаблон создан: {filepath}")


if __name__ == "__main__":
    create_template()