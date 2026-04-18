from openpyxl import load_workbook


class EquipmentCatalog:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.equipment_types = []
        self.room_categories = []
        self.position_coefficients = []

    def load(self):
        wb = load_workbook(self.filepath, data_only=True)

        self.equipment_types = self._read_sheet_as_dicts(wb["equipment_types"])
        self.room_categories = self._read_sheet_as_dicts(wb["room_categories"])
        self.position_coefficients = self._read_sheet_as_dicts(wb["position_coefficients"])

    @staticmethod
    def _read_sheet_as_dicts(ws):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(x).strip() if x is not None else "" for x in rows[0]]
        result = []

        for row in rows[1:]:
            if row is None:
                continue
            item = {}
            empty = True
            for i, value in enumerate(row):
                key = headers[i]
                if key:
                    item[key] = value
                    if value not in (None, ""):
                        empty = False
            if not empty:
                result.append(item)

        return result

    def get_equipment_type_names(self):
        return [row["type_name"] for row in self.equipment_types if row.get("type_name")]

    def get_equipment_type_by_id(self, type_id: str):
        for row in self.equipment_types:
            if str(row.get("type_id")) == str(type_id):
                return row
        return None

    def get_equipment_type_by_name(self, type_name: str):
        for row in self.equipment_types:
            if str(row.get("type_name")) == str(type_name):
                return row
        return None

    def get_room_categories(self):
        return [row["room_category"] for row in self.room_categories if row.get("room_category")]

    def get_room_category_defaults(self, room_category: str):
        for row in self.room_categories:
            if str(row.get("room_category")) == str(room_category):
                return row
        return None

    def get_position_names(self):
        return [row["position_name"] for row in self.position_coefficients if row.get("position_name")]

    def get_position_r(self, position_name: str):
        for row in self.position_coefficients:
            if str(row.get("position_name")) == str(position_name):
                return row.get("r_value")
        return None