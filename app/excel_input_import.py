from openpyxl import load_workbook

from app.schemas import ProjectInput


DEFAULT_CONSTANTS = {
    "k1": 0.5,
    "k2": 0.7,
    "k_empirical": 180.0,
    "z_m": 1.1,
    "ko": 0.8,
    "a": 1.25,
}


NOT_CALCULATE_TYPE_ID = "not_calculate"
NOT_CALCULATE_ROOM_CATEGORY = "Не рассчитывать"


def normalize_header(value) -> str:
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = text.replace("\n", " ")
    text = " ".join(text.split())
    return text


def find_column_indexes(header_row) -> dict:
    headers = {}
    for index, value in enumerate(header_row, start=1):
        name = normalize_header(value)

        if name in [
            "наименование помещения",
            "помещение",
            "название помещения",
            "room",
            "room_name",
        ]:
            headers["room_name"] = index

        elif name in [
            "оборудование",
            "наименование оборудования",
            "equipment",
            "equipment_name",
        ]:
            headers["equipment_name"] = index

        elif name in [
            "ширина",
            "ширина мм",
            "ширина, мм",
            "width",
            "width_mm",
        ]:
            headers["width_mm"] = index

        elif name in [
            "глубина",
            "глубина мм",
            "глубина, мм",
            "depth",
            "depth_mm",
        ]:
            headers["depth_mm"] = index

    required = ["room_name", "equipment_name", "width_mm", "depth_mm"]
    missing = [key for key in required if key not in headers]

    if missing:
        raise ValueError(
            "В Excel не найдены обязательные колонки: "
            + ", ".join(missing)
            + "\n\nНужны колонки: Наименование помещения, Оборудование, Ширина, Глубина."
        )

    return headers


def parse_float(value, row_number: int, column_name: str) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError(f"Строка {row_number}: пустое значение в колонке '{column_name}'.")

    text = str(value).strip().replace(",", ".")

    try:
        number = float(text)
    except Exception:
        raise ValueError(
            f"Строка {row_number}: значение '{value}' в колонке '{column_name}' не является числом."
        )

    if number < 0:
        raise ValueError(
            f"Строка {row_number}: значение в колонке '{column_name}' не может быть меньше 0."
        )

    return number


def import_customer_excel_to_project_input(filepath: str) -> ProjectInput:
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Excel-файл пустой.")

    headers = find_column_indexes(rows[0])

    rooms_map = {}

    for excel_row_index, row in enumerate(rows[1:], start=2):
        room_name_cell = row[headers["room_name"] - 1]
        equipment_name_cell = row[headers["equipment_name"] - 1]
        width_cell = row[headers["width_mm"] - 1]
        depth_cell = row[headers["depth_mm"] - 1]

        room_name = "" if room_name_cell is None else str(room_name_cell).strip()
        equipment_name = "" if equipment_name_cell is None else str(equipment_name_cell).strip()

        # Полностью пустые строки просто пропускаем
        if not room_name and not equipment_name and width_cell in (None, "") and depth_cell in (None, ""):
            continue

        if not room_name:
            raise ValueError(f"Строка {excel_row_index}: не заполнено наименование помещения.")

        if not equipment_name:
            raise ValueError(f"Строка {excel_row_index}: не заполнено наименование оборудования.")

        width_mm = parse_float(width_cell, excel_row_index, "Ширина")
        depth_mm = parse_float(depth_cell, excel_row_index, "Глубина")

        if room_name not in rooms_map:
            rooms_map[room_name] = {
                "room_name": room_name,
                "room_category": NOT_CALCULATE_ROOM_CATEGORY,
                "equipment": []
            }

        rooms_map[room_name]["equipment"].append({
            "name": equipment_name,
            "equipment_type_id": NOT_CALCULATE_TYPE_ID,
            "quantity": 1,
            "qy_kw": 0.0,
            "ka_w_per_kw": 0.0,
            "kz": 0.0,
            "position": "",
            "width_mm": width_mm,
            "depth_mm": depth_mm,
            "room_name": room_name,
        })

    if not rooms_map:
        raise ValueError("В Excel не найдено ни одной строки оборудования.")

    data = {
        "constants": DEFAULT_CONSTANTS,
        "rooms": list(rooms_map.values())
    }

    return ProjectInput(**data)